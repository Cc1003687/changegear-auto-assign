"""
ChangeGear 自動派單程式 v6
新增:
  - Requester's Item 使用正確的 popup 流程 (RadComboBox + AddImageButton + CloseLinkButton)
  - Accept 按鈕自動點擊 (Save 後)
  - 方案B: SQLite 歷史 Ticket 學習資料庫 (SequenceMatcher 相似度比對)
"""

import asyncio
import json
import logging
import os
import re
import sqlite3
import sys
from datetime import datetime, timedelta
from difflib import SequenceMatcher

import anthropic
import openpyxl
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from playwright.async_api import async_playwright, Page

EXCEL_PATH    = "ChangeGear_AutoAssign_Rules.xlsx"
DB_PATH       = "changegear_history.db"
CMDB_DB_PATH  = "cmdb_owners.db"     # Build CMDB DB.bat 產生的 CMDB owner 資料庫

# ── 縮短 Requester's Item 前綴，方便閱讀 ──────────────────
_RIRC = "DynamicLayoutControl1_ImpactedResourcesPIT_irc_18"
SEL = {
    "save":            "#ActionBarControl1_Save_Button",
    "accept":          "#ActionBarControl1_Accept_Button",
    "owner_input":     "#DynamicLayoutControl1_ItemOwner_PersonChooser_GridLookupPC_I",
    "assigned_input":  "#DynamicLayoutControl1_AssignedTo_PersonChooser_GridLookupPC_I",
    "impact":          "#DynamicLayoutControl1_Impact_ddl_10",
    "urgency":         "#DynamicLayoutControl1_Urgency_ddl_12",
    "priority":        "#DynamicLayoutControl1_Priority_ddl_14",
    "due_date":        "#DynamicLayoutControl1_DueDate_dtc_16_DueDate_Date_I",
    "inc_type_btn":    "[id*='IncidentRequestType'][id$='_B-1']",
    # Requester's Item popup（已驗證）
    # ri_open = dropDownPopupButton（A tag），開啟含搜尋 ComboBox 的 "Select Item" popup ✓
    # AddImpactedResouce_Button1（INPUT type=image）= Impact Analysis，不用此
    "ri_open":    f"#{_RIRC}_dropDownPopupButton",
    "ri_combo_id": f"{_RIRC}_ASPxPopupControl1_EntityChooserDD_EntityRadComboBox",
    "ri_arrow":   f"#{_RIRC}_ASPxPopupControl1_EntityChooserDD_EntityRadComboBox_Arrow",
    "ri_add":     f"#{_RIRC}_ASPxPopupControl1_AddImageButton",
    "ri_listbox": f"#{_RIRC}_ASPxPopupControl1_SelectedItemsLB",
    "ri_close":   f"#{_RIRC}_ASPxPopupControl1_CloseLinkButton",
}

IMPACT_MAP   = {"1 - Major":"1","2 - Significant":"2","3 - Minor":"3","4 - Routine":"4","4 - Low":"4"}
URGENCY_MAP  = {"1 - Emergency":"1","2 - High":"2","3 - Medium":"3","4 - Low":"4"}
PRIORITY_MAP = {"1 - Critical":"1","2 - High":"2","3 - Medium":"3","4 - Low":"4"}

# ── 有效 Incident Type 清單（來源：ITSM ticket types CSV，Revised 2025.Jul.1）──────
# 格式：(L1, L2, L3, Responsible)；L2/L3 為空字串代表該層不存在
VALID_INCIDENT_TYPES = [
    ("Complaint",       "",                       "",                         "All"),
    ("Error or Failure","Applications",           "",                         "Application"),
    ("Error or Failure","Infrastructure",         "",                         "Infrastructure"),
    ("Inquiry",         "",                       "",                         "All"),
    ("IT2IT Request",   "Applications",           "",                         "Application"),
    ("IT2IT Request",   "Infrastructure",         "",                         "infra/Helpdesk"),
    ("Service Request", "Applications",           "Standard Request",         "Application"),
    ("Service Request", "Applications",           "New Demand",               "Application"),
    ("Service Request", "Infrastructure",         "Standard Request",         "Helpdesk/Infra"),
    ("Service Request", "Infrastructure",         "Standard Request - nonSLA","Helpdesk/Infra"),
    ("Service Request", "Infrastructure",         "New Demand",               "Helpdesk/Infra"),
    ("Service Request", "Account Management",     "",                         "All"),
    ("Service Request", "Long lead time services","",                         "Helpdesk"),
    ("Service Request", "Onboarding/Offboarding", "",                         "All"),
]
# Responsible 欄位中含 helpdesk（含大小寫）即視為 Help Desk 工單
_HELPDESK_RESP = {"helpdesk", "helpdesk/infra", "infra/helpdesk"}


def is_helpdesk_owner(owner: str) -> bool:
    """判斷 owner 欄位是否為 Help Desk 負責。"""
    o = owner.strip().lower()
    return o in ("help desk", "helpdesk") or "help desk" in o or "helpdesk" in o


# ── Excel 規則載入 ──────────────────────────────────────
def load_rules(path: str) -> dict:
    wb = openpyxl.load_workbook(path)
    ws1 = wb["關鍵字派單規則"]
    keyword_rules = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if not row[0] or "請新增" in str(row[0]):
            continue
        kw, owner, assigned, inc_parent, inc_child, inc_item = (
            str(row[i]).strip() if row[i] else "" for i in range(6)
        )
        keyword_rules.append({"keyword":kw,"owner":owner,"assigned_to":assigned,
                               "inc_parent":inc_parent,"inc_child":inc_child,"inc_item":inc_item})

    ws2 = wb["Requester's Item 對應"]
    requester_items = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row[0] or "請新增" in str(row[0]):
            continue
        requester_items.append({"keyword":str(row[0]).strip(),"item":str(row[1]).strip()})

    ws4 = wb["程式設定"]
    config = {}
    for row in ws4.iter_rows(min_row=3, values_only=True):
        if row[0] and row[1]:
            config[str(row[0]).strip()] = str(row[1]).strip()

    return {
        "keyword_rules":    keyword_rules,
        "requester_items":  requester_items,
        "base_url":         config.get("系統網址","https://your-changegear-server.example.com/CGWeb"),
        "scan_interval":    int(config.get("掃描間隔（分鐘）","30")),
        "due_date_days":    int(config.get("Due Date 天數","5")),
        "log_path":         config.get("Log 檔案路徑","auto_assign.log"),
        "headless":         config.get("headless 模式","True").lower()=="true",
        "ad_account":       config.get("AD 帳號",""),
        "ad_password":      config.get("AD 密碼",""),
        "default_owner":    config.get("預設 Owner（無匹配時）","Help Desk"),
        "default_assigned": config.get("預設 Assigned To（無匹配）","IT-Helpdesk"),
        "db_similarity":    float(config.get("歷史比對相似度門檻","0.65")),
        "claude_api_key":   config.get("Claude API Key",""),
        "claude_model":     config.get("Claude 模型","claude-haiku-4-5"),
        "claude_candidates":int(config.get("Claude 候選工單數","8")),
    }

RULES = load_rules(EXCEL_PATH)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(RULES["log_path"], encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════
# 方案B: SQLite 歷史學習資料庫
# ══════════════════════════════════════════════════════════
def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS assignments (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_id    TEXT UNIQUE,
            oid          TEXT,
            summary      TEXT,
            description  TEXT,
            requester    TEXT,
            owner        TEXT,
            assigned_to  TEXT,
            inc_parent   TEXT,
            inc_child    TEXT,
            inc_item     TEXT,
            req_item     TEXT,
            bot_assigned INTEGER DEFAULT 0,
            corrected    INTEGER DEFAULT 0,
            created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # 舊版 DB 補欄位（PRAGMA 確認不存在才 ALTER，避免 exception 被吞）
    existing_cols = {r[1] for r in conn.execute("PRAGMA table_info(assignments)").fetchall()}
    for col, defn in [
        ("oid",                  "TEXT DEFAULT ''"),
        ("requester",            "TEXT DEFAULT ''"),
        ("bot_assigned",         "INTEGER DEFAULT 0"),
        ("corrected",            "INTEGER DEFAULT 0"),
        ("updated_at",           "TIMESTAMP DEFAULT ''"),
        # ── stats.py 用：bot 原本派的值（被 correction_scan 覆蓋前的快照）──
        ("original_owner",       "TEXT DEFAULT ''"),
        ("original_assigned_to", "TEXT DEFAULT ''"),
        ("original_inc_parent",  "TEXT DEFAULT ''"),
        ("original_inc_child",   "TEXT DEFAULT ''"),
        ("original_inc_item",    "TEXT DEFAULT ''"),
        ("original_req_item",    "TEXT DEFAULT ''"),
        # ── stats.py 用：決策來源與信心度 ──
        ("decision_source",      "TEXT DEFAULT ''"),
        ("confidence",           "REAL DEFAULT 0.0"),
    ]:
        if col not in existing_cols:
            conn.execute(f"ALTER TABLE assignments ADD COLUMN {col} {defn}")
            log.info(f"DB 補欄位: {col}")
    conn.commit()
    conn.close()
    log.info(f"SQLite 歷史資料庫已初始化: {DB_PATH}")

def db_track_ticket(ticket_id: str, oid: str, summary: str,
                    description: str, requester: str):
    """追蹤非 Help Desk 工單（或空白工單），供修正學習使用。

    bot_assigned 狀態說明：
      0 = 歷史人工派單（高信任）
      1 = bot 自動派單（待驗證）
      2 = 追蹤中，未派單（偵測到但不是 Help Desk；等下次掃描看是否有人工填入）

    僅在 ticket_id 不存在時才寫入（避免覆蓋已有資料）。
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute("""
            INSERT OR IGNORE INTO assignments
            (ticket_id, oid, summary, description, requester,
             owner, assigned_to, inc_parent, inc_child, inc_item, req_item,
             bot_assigned, corrected, updated_at)
            VALUES (?, ?, ?, ?, ?, '', '', '', '', '', '', 2, 0, CURRENT_TIMESTAMP)
        """, (ticket_id, oid, summary, (description or "")[:500], requester))
        conn.commit()
        conn.close()
        log.debug(f"追蹤記錄建立: {ticket_id}")
    except Exception as e:
        log.warning(f"db_track_ticket 失敗: {e}")


def db_save(ticket_id: str, oid: str, summary: str, description: str,
            requester: str, a: dict):
    """儲存 bot 派單記錄到 DB。
    bot_assigned=1 代表此筆由 bot 寫入，可供修正掃描使用。
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        # _source 從 determine_assignment 帶過來的字串裡可能含 "Claude(conf=0.92,src=cmdb)"
        # 抽取 src= 後面那個詞當 decision_source
        src_str = a.get("_source", "")
        m_src   = re.search(r"src=([a-zA-Z]+)", src_str)
        decision_source = m_src.group(1) if m_src else (
            "cmdb" if "CMDB" in src_str else
            "claude" if "Claude" in src_str else
            "db" if "DB:" in src_str else
            "excel" if "Excel" in src_str else
            "default" if "預設" in src_str else
            ""
        )
        confidence = float(a.get("_score", 0.0) or 0.0)

        conn.execute("""
            INSERT OR REPLACE INTO assignments
            (ticket_id, oid, summary, description, requester,
             owner, assigned_to, inc_parent, inc_child, inc_item, req_item,
             original_owner, original_assigned_to,
             original_inc_parent, original_inc_child, original_inc_item, original_req_item,
             decision_source, confidence,
             bot_assigned, corrected, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?,
                    ?, ?,
                    1, 0, CURRENT_TIMESTAMP)
        """, (
            ticket_id, oid, summary, description[:500], requester,
            a["owner"], a["assigned_to"],
            a["inc_parent"], a["inc_child"], a["inc_item"], a.get("requester_item", ""),
            # original_* 與 owner/inc_* 同步寫入，作為 bot 原始選擇的快照
            a["owner"], a["assigned_to"],
            a["inc_parent"], a["inc_child"], a["inc_item"], a.get("requester_item", ""),
            decision_source, confidence,
        ))
        conn.commit()
        conn.close()
        log.debug(f"已儲存歷史記錄: {ticket_id} (source={decision_source}, conf={confidence:.2f})")
    except Exception as e:
        log.warning(f"DB 儲存失敗: {e}")

def db_find_similar(summary: str, description: str = "", requester: str = "") -> dict | None:
    """從歷史 DB 找最相似的派單記錄。

    加權分數組成：
      ① 文字相似度 (summary + description)        基礎分
      ② Requester 相符（寄件者相同）               +0.30 / 部分 +0.15
      ③ 來源可信度加權（trust_bonus）：
           corrected   = 1  → 人工修正過，最可信   × 1.20
           bot_assigned= 0  → 歷史人工派單，可信   × 1.10
           bot_assigned= 1, corrected=0 → bot 派未驗證，不加分 × 1.00
      最終 score = min(1.0, text_ratio * 0.70 + req_bonus) * trust_bonus
      門檻：RULES["db_similarity"]（預設 0.65）
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute(
            "SELECT ticket_id, summary, description, requester, owner, assigned_to, "
            "inc_parent, inc_child, inc_item, req_item, bot_assigned, corrected "
            "FROM assignments "
            "ORDER BY created_at DESC LIMIT 500"
        ).fetchall()
        conn.close()

        if not rows:
            return None

        query_text = f"{summary} {description}".lower().strip()
        req_lower  = requester.lower().strip()
        best_score = 0.0
        best_row   = None

        for row in rows:
            # ① 文字相似度
            hist_text  = f"{row[1]} {row[2] or ''}".lower().strip()
            text_ratio = SequenceMatcher(None, query_text, hist_text).ratio()

            # ② Requester 加分
            req_bonus = 0.0
            if req_lower and row[3]:
                hist_req = row[3].lower().strip()
                if req_lower == hist_req:
                    req_bonus = 0.30
                elif req_lower in hist_req or hist_req in req_lower:
                    req_bonus = 0.15

            # ③ 來源可信度加權
            bot_assigned = row[10]
            corrected    = row[11]
            if corrected:
                trust_bonus = 1.20      # 人工修正過，最可信
            elif not bot_assigned:
                trust_bonus = 1.10      # 歷史人工派單，可信
            else:
                trust_bonus = 1.00      # bot 派但未被驗證，不額外加分

            score = min(1.0, (text_ratio * 0.70 + req_bonus) * trust_bonus)

            if score > best_score:
                best_score = score
                best_row   = row

        threshold = RULES["db_similarity"]
        if best_score >= threshold and best_row:
            trust_label = (
                "人工修正" if best_row[11] else
                "歷史人工" if not best_row[10] else
                "bot派單"
            )
            log.info(f"歷史比對命中 [{trust_label}] (score={best_score:.2f}): {best_row[0]}")
            return {
                "owner":          best_row[4],
                "assigned_to":    best_row[5],
                "inc_parent":     best_row[6],
                "inc_child":      best_row[7],
                "inc_item":       best_row[8],
                "requester_item": best_row[9] or "",
                "_source":        f"DB:{best_row[0]}[{trust_label}](score={best_score:.2f})",
                "_score":         best_score,   # 供 CMDB 審查層使用
            }
        return None
    except Exception as e:
        log.warning(f"DB 查詢失敗: {e}")
        return None


# ══════════════════════════════════════════════════════════
# Claude AI 派單輔助
# ══════════════════════════════════════════════════════════
def db_get_candidates(summary: str, description: str = "", requester: str = "",
                      top_n: int = 8) -> list[dict]:
    """取得歷史 DB 中相似度最高的前 N 筆作為 Claude 參考（無門檻，只取 top N）。"""
    try:
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute(
            "SELECT ticket_id, summary, description, requester, owner, assigned_to, "
            "inc_parent, inc_child, inc_item, req_item, bot_assigned, corrected "
            "FROM assignments ORDER BY created_at DESC LIMIT 300"
        ).fetchall()
        conn.close()

        query_text = f"{summary} {description}".lower().strip()
        req_lower  = requester.lower().strip()
        scored = []

        for row in rows:
            hist_text  = f"{row[1]} {row[2] or ''}".lower().strip()
            text_ratio = SequenceMatcher(None, query_text, hist_text).ratio()
            req_bonus  = 0.0
            if req_lower and row[3]:
                hist_req = row[3].lower().strip()
                if req_lower == hist_req:
                    req_bonus = 0.30
                elif req_lower in hist_req or hist_req in req_lower:
                    req_bonus = 0.15
            trust = 1.20 if row[11] else (1.10 if not row[10] else 1.00)
            score = min(1.0, (text_ratio * 0.70 + req_bonus) * trust)
            scored.append((score, row))

        scored.sort(key=lambda x: x[0], reverse=True)
        result = []
        for score, row in scored[:top_n]:
            result.append({
                "ticket_id":   row[0],
                "summary":     row[1],
                "requester":   row[3] or "",
                "owner":       row[4],
                "assigned_to": row[5],
                "inc_parent":  row[6],
                "inc_child":   row[7],
                "inc_item":    row[8],
                "req_item":    row[9] or "",
                "source":      "人工修正" if row[11] else ("歷史人工" if not row[10] else "bot派單"),
                "score":       round(score, 3),
            })
        return result
    except Exception as e:
        log.warning(f"db_get_candidates 失敗: {e}")
        return []


def db_get_cmdb_critical_names() -> list[str]:
    """
    取得 CMDB DB 中所有 critical_name（資產名稱），提供給 Claude 作為
    Requester's Item 的候選清單。

    Claude 會從這個清單裡挑選最符合工單內容的 CI，系統再用該 CI 從
    CMDB 反查擁有者（owner / team_owner）作為實際派單對象。
    """
    try:
        conn = sqlite3.connect(CMDB_DB_PATH)
        rows = conn.execute(
            "SELECT critical_name FROM cmdb_items "
            "WHERE critical_name IS NOT NULL AND critical_name != '' "
            "ORDER BY critical_name"
        ).fetchall()
        conn.close()
        return [r[0] for r in rows]
    except Exception as e:
        log.debug(f"db_get_cmdb_critical_names 失敗: {e}")
        return []


def db_get_req_items() -> list[str]:
    """取得 DB 中所有不重複的 Requester's Item，提供 Claude 選擇範圍。"""
    try:
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute(
            "SELECT DISTINCT req_item FROM assignments "
            "WHERE req_item IS NOT NULL AND req_item != '' ORDER BY req_item"
        ).fetchall()
        conn.close()
        return [r[0] for r in rows]
    except Exception:
        return []


def db_get_incident_types() -> list[str]:
    """回傳 CSV 定義的有效 Incident Type 清單（不查 DB，確保分類正確不失真）。"""
    result = []
    for l1, l2, l3, _ in VALID_INCIDENT_TYPES:
        if l3:
            result.append(f"{l1} > {l2} > {l3}")
        elif l2:
            result.append(f"{l1} > {l2}")
        else:
            result.append(l1)
    return result


# ══════════════════════════════════════════════════════════
# Due Date 工作天計算
# ══════════════════════════════════════════════════════════
def calc_business_due_date(base_date: datetime, business_days: int) -> datetime:
    """
    計算 base_date 加上 business_days 個工作天的日期（略過週六日）。

    例：今天週三 + 4 工作天 → 下週二（跳過週六日）
    """
    due = base_date
    added = 0
    while added < business_days:
        due += timedelta(days=1)
        if due.weekday() < 5:   # 0=Mon … 4=Fri
            added += 1
    return due


# ══════════════════════════════════════════════════════════
# CMDB 鑑別審查層
# ══════════════════════════════════════════════════════════
def cmdb_lookup(req_item: str) -> dict | None:
    """
    從 CMDB DB（cmdb_owners.db）查找與 req_item（Requester's Item）
    最相符的 CMDB 項目，回傳 owner / team_owner 等欄位。

    查找順序：
      1. critical_name 精確匹配
      2. SequenceMatcher 模糊匹配（ratio >= 0.65 才採用）
    找不到 CMDB DB 或無符合項目時回傳 None（不影響主流程）。
    """
    if not req_item:
        return None
    try:
        conn = sqlite3.connect(CMDB_DB_PATH)
        # 精確匹配
        row = conn.execute(
            "SELECT critical_name, owner, co_owner, tech_owner, team_owner "
            "FROM cmdb_items WHERE critical_name = ? LIMIT 1",
            (req_item,)
        ).fetchone()
        if not row:
            # 模糊匹配：全表掃描取最高相似度
            all_rows = conn.execute(
                "SELECT critical_name, owner, co_owner, tech_owner, team_owner FROM cmdb_items"
            ).fetchall()
            best_ratio, best_row = 0.0, None
            for r in all_rows:
                ratio = SequenceMatcher(None, req_item.lower(),
                                        (r[0] or "").lower()).ratio()
                if ratio > best_ratio:
                    best_ratio, best_row = ratio, r
            if best_ratio >= 0.65:
                row = best_row
        conn.close()
        if row:
            return {
                "critical_name": row[0] or "",
                "owner":         row[1] or "",
                "co_owner":      row[2] or "",
                "tech_owner":    row[3] or "",
                "team_owner":    row[4] or "",
            }
        return None
    except Exception as e:
        log.debug(f"CMDB 查詢失敗（DB 可能尚未建立）: {e}")
        return None


def _name_match(a: str, b: str) -> bool:
    """寬鬆名字比對：兩者之一包含另一者，或 SequenceMatcher >= 0.75。"""
    a, b = a.strip().lower(), b.strip().lower()
    if not a or not b:
        return False   # 任一空值視為「無法比對」→ 不通過
    return (a in b or b in a or
            SequenceMatcher(None, a, b).ratio() >= 0.75)


def cmdb_validate(assigned_to: str, owner: str, req_item: str) -> tuple[bool, str]:
    """
    CMDB 鑑別審查（DB 比對信心 >= 0.75 才呼叫）：
      ① assigned_to（歷史比對派單人）↔ CMDB.owner（CI 實際負責人）
      ② owner（歷史比對負責群組）    ↔ CMDB.team_owner（CI 負責團隊）

    回傳：
      (True,  說明文字) → 一致，通過審查
      (False, 不一致原因) → 不一致，需進入二次判定
    """
    cmdb = cmdb_lookup(req_item)
    if not cmdb:
        return True, "CMDB 無對應記錄，略過審查"

    assigned_ok = _name_match(assigned_to, cmdb["owner"])
    team_ok     = _name_match(owner,       cmdb["team_owner"])

    if assigned_ok and team_ok:
        return True, (
            f"CMDB 審查通過 | CI={cmdb['critical_name']} "
            f"owner={cmdb['owner']} team={cmdb['team_owner']}"
        )

    reasons = []
    if not assigned_ok:
        reasons.append(
            f"Assigned To「{assigned_to}」≠ CMDB.owner「{cmdb['owner']}」"
        )
    if not team_ok:
        reasons.append(
            f"Owner「{owner}」≠ CMDB.team_owner「{cmdb['team_owner']}」"
        )
    return False, "；".join(reasons)


def cmdb_direct_lookup(req_item: str) -> dict | None:
    """
    依 Requester's Item 名稱「直接」從 CMDB 取得擁有者，作為派單候選。

    與 cmdb_validate 不同：
      - cmdb_validate 是審查層（驗證既有派單建議是否正確）
      - cmdb_direct_lookup 是來源層（產生派單建議）

    對應關係：
      CMDB.owner       → 派單 Assigned To（個人擁有者）
      CMDB.team_owner  → 派單 Owner       （團隊擁有者）

    回傳 dict（owner / assigned_to / requester_item / _score / _source），
    或 None（CMDB 無對應、或缺少必要的擁有者資料）。

    _score 設為 0.80（介於 CMDB 審查門檻 0.75 與 AI 門檻 0.85 之間），
    觸發「信心 >= 0.75 跳過 Help Desk 過濾」例外規則。
    """
    if not req_item:
        return None
    cmdb = cmdb_lookup(req_item)
    if not cmdb:
        return None

    cmdb_assigned = (cmdb.get("owner") or "").strip()
    cmdb_team     = (cmdb.get("team_owner") or "").strip()

    if not cmdb_assigned:
        # 沒個人擁有者 → 不派單（無法決定 Assigned To）
        return None
    if not cmdb_team:
        # 缺團隊資料 → 預設 Help Desk
        cmdb_team = RULES.get("default_owner", "Help Desk")

    return {
        "owner":          cmdb_team,
        "assigned_to":    cmdb_assigned,
        "requester_item": cmdb.get("critical_name", ""),
        "inc_parent":     "",   # CMDB 無 Incident Type 資訊，留給後續層級補
        "inc_child":      "",
        "inc_item":       "",
        "_score":         0.80,
        "_source":        f"CMDB直接派單(CI={cmdb.get('critical_name','')})",
    }


def extract_hi_name(description: str) -> str:
    """
    從工單描述中提取問候語（Hi / Hello / Dear + 名字）裡的收件人姓名。

    匹配格式示例：
      "Hi John,"  → "John"
      "Dear Mary Smith," → "Mary Smith"
      "Hello IT-Leo.Chen," → "IT-Leo.Chen"

    常見非名詞（all / team / everyone …）會被過濾掉。
    """
    if not description:
        return ""
    patterns = [
        # 「Hi/Hello/Dear 名字,」或「Hi 名字\n」
        r"(?:Hi|Hello|Dear)[,\s]+([A-Za-z][A-Za-z0-9\-\.]{1,30}"
        r"(?:\s+[A-Za-z][A-Za-z0-9\-\.]{1,20})?)\s*[,\.\n\r!]",
        # 結尾型：「Hi 名字」（句尾無標點）
        r"(?:Hi|Hello|Dear)[,\s]+([A-Za-z][A-Za-z0-9\-\.]{1,30})\s*$",
    ]
    stopwords = {
        "all", "team", "everyone", "there", "sir", "madam",
        "folks", "it", "support", "helpdesk", "help",
    }
    for pat in patterns:
        m = re.search(pat, description, re.IGNORECASE | re.MULTILINE)
        if m:
            name = m.group(1).strip()
            if name.lower() not in stopwords and len(name) >= 2:
                return name
    return ""


def db_init_feedback_table():
    """建立 feedback 表（人工教學紀錄）。
    每次呼叫都是 IF NOT EXISTS，安全可重複呼叫。
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS feedback (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id           TEXT,
                summary             TEXT,
                description         TEXT,
                requester           TEXT,
                req_item            TEXT,
                wrong_owner         TEXT,
                wrong_assigned_to   TEXT,
                correct_owner       TEXT,
                correct_assigned_to TEXT,
                reason              TEXT,
                created_at          TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
    finally:
        conn.close()


def db_get_relevant_feedback(summary: str, description: str = "",
                             req_item: str = "",
                             top_n: int = 5,
                             min_ratio: float = 0.30) -> list[dict]:
    """根據新工單內容，從 feedback 表挑出 top_n 個最相關的歷史教訓。

    比對策略：對每筆 feedback 計算 SequenceMatcher 相似度
              = (summary + description + req_item) vs (feedback 的同樣三欄)
    回傳已依相似度由高到低排序的清單；低於 min_ratio 的略過。

    供 claude_assign() 注入到 user prompt，讓 Claude 在判斷時參考。
    """
    try:
        db_init_feedback_table()   # 確保表存在
        conn = sqlite3.connect(DB_PATH)
        try:
            rows = conn.execute("""
                SELECT ticket_id, summary, description, requester, req_item,
                       wrong_owner, wrong_assigned_to,
                       correct_owner, correct_assigned_to, reason, created_at
                FROM feedback
                ORDER BY created_at DESC
            """).fetchall()
        finally:
            conn.close()

        if not rows:
            return []

        query_text = f"{summary} {description} {req_item}".lower().strip()
        scored = []
        for r in rows:
            fb_text = f"{r[1] or ''} {r[2] or ''} {r[4] or ''}".lower().strip()
            ratio = SequenceMatcher(None, query_text, fb_text).ratio() if fb_text else 0.0
            scored.append((ratio, r))

        scored.sort(key=lambda x: x[0], reverse=True)
        out = []
        for ratio, r in scored[:top_n]:
            if ratio < min_ratio:
                continue
            out.append({
                "ticket_id":           r[0],
                "summary":             r[1] or "",
                "req_item":            r[4] or "",
                "wrong_owner":         r[5] or "",
                "wrong_assigned_to":   r[6] or "",
                "correct_owner":       r[7] or "",
                "correct_assigned_to": r[8] or "",
                "reason":              r[9] or "",
                "similarity":          round(ratio, 2),
            })
        return out
    except Exception as e:
        log.warning(f"db_get_relevant_feedback 失敗: {e}")
        return []


def db_find_by_name(name: str,
                    summary: str = "", description: str = "") -> dict | None:
    """
    從歷史 DB 找 assigned_to 含有 name 的記錄，結合文字相似度計算信心分數。
    信心分數 < 0.85 時回傳 None（二次判定門檻較嚴）。

    分數組成：
      name_score（名字匹配程度）× 0.60 +
      text_ratio（標題+描述相似度）× 0.40
      × trust_bonus（來源可信度）
    """
    if not name:
        return None
    try:
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute(
            "SELECT ticket_id, summary, description, requester, owner, assigned_to, "
            "inc_parent, inc_child, inc_item, req_item, bot_assigned, corrected "
            "FROM assignments "
            "WHERE lower(assigned_to) LIKE ? "
            "ORDER BY corrected DESC, bot_assigned ASC, created_at DESC LIMIT 100",
            (f"%{name.lower()}%",)
        ).fetchall()
        conn.close()

        if not rows:
            return None

        query_text = f"{summary} {description}".lower().strip()
        best_score, best_row = 0.0, None

        for row in rows:
            assigned_lower = (row[5] or "").lower()
            name_lower     = name.lower()
            if name_lower == assigned_lower:
                name_score = 1.00
            elif name_lower in assigned_lower:
                name_score = 0.90
            else:
                name_score = 0.80

            # 文字相似度（有 query_text 才算，否則給中性分）
            if query_text:
                hist_text  = f"{row[1]} {row[2] or ''}".lower()
                text_ratio = SequenceMatcher(None, query_text, hist_text).ratio()
            else:
                text_ratio = 0.50

            # 來源可信度加權
            trust = 1.20 if row[11] else (1.10 if not row[10] else 1.00)
            score = min(1.0, (name_score * 0.60 + text_ratio * 0.40) * trust)

            if score > best_score:
                best_score, best_row = score, row

        if best_score < 0.85 or best_row is None:
            log.debug(
                f"db_find_by_name 信心不足: name={name} score={best_score:.2f} < 0.85"
            )
            return None

        trust_label = (
            "人工修正" if best_row[11] else
            "歷史人工" if not best_row[10] else "bot派單"
        )
        log.info(
            f"二次人名比對命中 [{trust_label}] (score={best_score:.2f}): "
            f"名字={name} → {best_row[5]}"
        )
        return {
            "owner":          best_row[4],
            "assigned_to":    best_row[5],
            "inc_parent":     best_row[6],
            "inc_child":      best_row[7],
            "inc_item":       best_row[8],
            "requester_item": best_row[9] or "",
            "_source":        f"DB人名({name})[{trust_label}](score={best_score:.2f})",
            "_score":         best_score,
        }
    except Exception as e:
        log.warning(f"db_find_by_name 失敗: {e}")
        return None


async def claude_assign(summary: str, description: str, requester: str,
                        candidates: list[dict]) -> dict | None:
    """呼叫 Claude API，根據新工單與候選歷史記錄決定最佳派單。

    回傳格式與 determine_assignment 相同的 dict，失敗時回傳 None。
    """
    api_key = RULES.get("claude_api_key", "")
    if not api_key or not candidates:
        return None

    # ── 組候選工單說明（含 CMDB 擁有者交叉比對）──────────────────────
    # 三信號合一：每個歷史候選的 req_item 都即時查 CMDB 註記實際擁有者，
    # Claude 可以直接看到「歷史派給 X，但 CMDB 顯示這資產屬於 Y」的衝突，
    # 不需多階段 fallback 也能在單次判斷中做出最佳決策。
    cand_lines = []
    for i, c in enumerate(candidates, 1):
        # 註記該候選 req_item 在 CMDB 中的擁有者（若有）
        cmdb_note = ""
        if c.get("req_item"):
            cmdb_hit = cmdb_lookup(c["req_item"])
            if cmdb_hit:
                cmdb_own  = (cmdb_hit.get("owner") or "").strip() or "(無)"
                cmdb_team = (cmdb_hit.get("team_owner") or "").strip() or "(無)"
                # 一致 / 不一致用符號標記，方便 Claude 識別衝突
                same = (cmdb_own.lower() in c["assigned_to"].lower()
                        or c["assigned_to"].lower() in cmdb_own.lower())
                tag  = "✓ 一致" if same else "✗ 不一致"
                cmdb_note = (
                    f"\n    CMDB 比對: CI={cmdb_hit['critical_name']} "
                    f"→ Owner(team)={cmdb_team} / Owner(person)={cmdb_own} [{tag}]"
                )
            else:
                cmdb_note = f"\n    CMDB 比對: 「{c['req_item']}」未在 CMDB 中"
        cand_lines.append(
            f"[{i}] 工單:{c['ticket_id']} 相似度:{c['score']} 來源:{c['source']}\n"
            f"    標題: {c['summary']}\n"
            f"    寄件者: {c['requester']}\n"
            f"    Owner: {c['owner']} / Assigned To: {c['assigned_to']}\n"
            f"    Incident Type: {c['inc_parent']} > {c['inc_child']} > {c['inc_item']}\n"
            f"    Requester's Item: {c['req_item']}{cmdb_note}"
        )
    candidates_text = "\n".join(cand_lines)

    # ── 可選的 Requester's Item 清單（來源：CMDB DB）─────────────────
    # 設計變更：清單來源從歷史 DB 改成 CMDB，讓 Claude 直接從實際資產
    # 清單中挑選最符合工單內容的 CI。系統會依 Claude 選定的 CI 從
    # CMDB 反查擁有者（owner / team_owner）作為實際派單對象。
    cmdb_items = db_get_cmdb_critical_names()
    if cmdb_items:
        req_items_text = "\n".join(f"- {r}" for r in cmdb_items)
    else:
        # CMDB 尚未建立時退回歷史 DB（向下相容）
        req_items = db_get_req_items()
        req_items_text = "\n".join(f"- {r}" for r in req_items) if req_items else "（無記錄）"

    # 有效的 Incident Type 清單（只能從這裡選，不可自行創造）
    inc_types = db_get_incident_types()
    inc_types_text = "\n".join(f"- {t}" for t in inc_types) if inc_types else "（DB 尚無記錄）"

    # ══ Prompt 結構（為了啟用 Prompt Caching，穩定內容放 system，動態內容放 user）══
    # 系統指令（穩定，可快取）：角色、規則、有效清單、JSON 格式
    #   - 多輪呼叫之間 inc_types / req_items 與規則不變 → 命中快取，成本 ≈ 0.1×
    #   - cache_control 標記放在「最後一個穩定 block」，標記前的所有 system 一起進快取

    # ── 載入 reflect.py 每週歸納的派單原則（深度學習）──────────────
    # 若 learned_principles.md 存在且有實質內容，整份注入 system prompt 最頂端，
    # 優先級高於人工教學紀錄。這份檔案是 Claude 從多筆 feedback/correction
    # 歸納出的「高層派單規則」，避免 bot 重複犯類似錯誤。
    principles_text = ""
    if os.path.exists("learned_principles.md"):
        try:
            with open("learned_principles.md", encoding="utf-8") as _pf:
                _content = _pf.read().strip()
            # 排除「無足夠案例可歸納」這種空殼檔
            if _content and "無足夠案例" not in _content and "## 原則" in _content:
                principles_text = (
                    f"\n## 已歸納原則（最高優先級，務必遵循）\n"
                    f"以下為過去人工修正案例的歸納結果，反映本組織的派單慣例。\n"
                    f"若新工單符合任一條原則，必須依該原則派單。\n\n"
                    f"{_content}\n\n"
                    f"════════════════════════════════════════════════════════\n"
                )
        except Exception as _e:
            log.debug(f"載入 learned_principles.md 失敗: {_e}")

    system_text = f"""你是 IT 服務台的派單助理。請根據新工單資訊與歷史參考工單，判斷最合適的派單結果。
{principles_text}

## 系統中有效的 Incident Type 清單（格式：第一層 > 第二層 > 第三層）
{inc_types_text}

## CMDB 資產清單（Requester's Item 必須從此清單中挑選）
{req_items_text}

## 0. 人工教學紀錄（最高優先順序）
若 user 訊息中含有「## ⚠ 人工教學紀錄」段落，請務必依下列原則處理：
  - 該段落中的每筆 Lesson 是過去同類工單被 IT 主管手動修正的紀錄
  - 若新工單與某筆 Lesson 相似（相似度高、req_item 相同、或工單性質一致），
    必須採用該 Lesson 的「正確派給」結果，並在 reasoning 中註明參考了哪個 Lesson
  - Lesson 的權重高於下方三信號（CMDB / 歷史 / 內容）
  - decision_source 在此情況請填 "feedback"

## 三信號決策原則（次重要）
你的判斷會綜合三個信號：
  ① 工單內容（標題、內文、寄件者）
  ② 歷史相似工單（candidates 區段，每筆含過去的 owner/assigned_to）
  ③ CMDB 資產擁有者（每個歷史候選旁邊都註記了 CMDB 比對結果）

權衡優先順序：
  - 若工單明確提到某項資產 + CMDB 有該資產 → 以 CMDB 擁有者為主
  - 若歷史候選顯示「CMDB 比對：一致」→ 可放心採用該歷史的 assigned_to
  - 若歷史候選顯示「CMDB 比對：不一致」→ 通常 CMDB 較可信（資產異動後人員變更），優先選 CMDB 擁有者
  - 若 CMDB 完全沒對應 → 採用歷史最相似工單的 assigned_to
  - 「人工修正」候選的權重最高，其次「歷史人工」，最後「bot派單」
  - 寄件者相同的歷史候選有更高參考價值

## 欄位規則
1. inc_parent / inc_child / inc_item 必須完整照抄上方「有效的 Incident Type 清單」中的某一筆，不可自行創造或修改名稱。
2. assigned_to 不可留空。優先採用「CMDB 比對一致」的歷史候選對應人員；若無，採用 CMDB 反查擁有者；若 CMDB 也沒有，採用最相似的歷史候選人員。
3. req_item 必須從上方「CMDB 資產清單」中挑選最符合工單描述的資產名稱。若工單描述模糊或無對應資產，留空字串。
4. 系統會根據你選的 req_item 從 CMDB 反查擁有者並覆蓋你回傳的 owner/assigned_to，因此 req_item 選對非常重要。
5. 只回傳 JSON，不要任何其他文字或 markdown。

## 回傳格式
{{
  "owner": "負責人",
  "assigned_to": "指派人員（不可空白）",
  "inc_parent": "Incident Type 第一層（照抄清單）",
  "inc_child": "Incident Type 第二層（照抄清單）",
  "inc_item": "Incident Type 第三層（照抄清單）",
  "req_item": "從 CMDB 資產清單中挑選的 CI 名稱（或空字串）",
  "decision_source": "最關鍵來源: 'feedback' | 'cmdb' | 'history' | 'content' | 'hybrid'",
  "confidence": 0.0至1.0,
  "reasoning": "一句話說明為何選此 assigned_to（例：CMDB 與歷史一致 / CMDB 覆蓋歷史 / 內容明指）"
}}"""

    # ── 抓取人工教學紀錄（feedback）中最相關的條目 ──────────────────
    # 使用者透過 teach.py 寫入的「應該派給誰、為什麼」會被存到 feedback 表，
    # 這裡根據新工單內容挑出相關度最高的 5 筆注入 prompt，
    # 讓 Claude 在判斷時優先參考過去的人工修正。
    feedback_items = db_get_relevant_feedback(summary, description, "")
    feedback_text  = ""
    if feedback_items:
        log.info(f"  注入 {len(feedback_items)} 筆人工教學紀錄到 prompt")
        fb_lines = ["## ⚠ 人工教學紀錄（過去類似工單的修正，務必參考）"]
        for i, f in enumerate(feedback_items, 1):
            fb_lines.append(
                f"[Lesson {i}] 過去工單 {f['ticket_id']}（與本工單相似度 {f['similarity']}）\n"
                f"    工單摘要: {f['summary'][:80]}\n"
                f"    Requester's Item: {f['req_item']}\n"
                f"    ✗ 錯誤派給: Owner={f['wrong_owner']} / Assigned={f['wrong_assigned_to']}\n"
                f"    ✓ 正確派給: Owner={f['correct_owner']} / Assigned={f['correct_assigned_to']}\n"
                f"    判斷依據: {f['reason']}"
            )
        feedback_text = "\n".join(fb_lines) + "\n\n"

    # 使用者訊息（動態，每次不同 → 不快取）：教學紀錄 + 新工單 + 候選清單
    user_text = f"""{feedback_text}## 新工單
- 標題（Summary）: {summary}
- 內文（Description）: {description[:600] if description else "（無）"}
- 寄件者（Requester）: {requester}

## 歷史參考工單（依相似度排序）
{candidates_text}"""

    try:
        client = anthropic.AsyncAnthropic(api_key=api_key)
        response = await client.messages.create(
            model=RULES.get("claude_model", "claude-haiku-4-5"),
            max_tokens=512,
            system=[
                {
                    "type": "text",
                    "text": system_text,
                    "cache_control": {"type": "ephemeral"},   # 5 分鐘 TTL
                }
            ],
            messages=[{"role": "user", "content": user_text}],
        )

        # ── 觀測快取命中率 ──────────────────────────────────────────
        u = response.usage
        cache_create = getattr(u, "cache_creation_input_tokens", 0) or 0
        cache_read   = getattr(u, "cache_read_input_tokens", 0) or 0
        if cache_create or cache_read:
            log.info(
                f"[Claude cache] read={cache_read} write={cache_create} "
                f"uncached={u.input_tokens} output={u.output_tokens}"
            )

        raw = response.content[0].text.strip()

        # 移除可能的 markdown code block
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)

        data = json.loads(raw)
        confidence = float(data.get("confidence", 0))
        decision_src = str(data.get("decision_source", "unknown")).strip()
        log.info(
            f"[Claude] confidence={confidence:.2f} source={decision_src} | "
            f"{data.get('reasoning','')[:80]}"
        )

        if confidence < 0.85:
            log.warning(f"Claude 信心不足 ({confidence:.2f} < 0.85)，降回傳統邏輯")
            return None

        assigned_to = str(data.get("assigned_to", "")).strip()
        # Claude 回傳空白時，從候選工單中取第一筆有值的 assigned_to 補入
        if not assigned_to:
            for c in candidates:
                if c.get("assigned_to"):
                    assigned_to = c["assigned_to"]
                    log.warning(f"Claude assigned_to 空白，補入候選值: {assigned_to}")
                    break

        return {
            "owner":            str(data.get("owner", "")),
            "assigned_to":      assigned_to,
            "inc_parent":       str(data.get("inc_parent", "")),
            "inc_child":        str(data.get("inc_child", "")),
            "inc_item":         str(data.get("inc_item", "")),
            "requester_item":   str(data.get("req_item", "")),
            "decision_source":  decision_src,
            "confidence":       confidence,
            "_source":          f"Claude(conf={confidence:.2f},src={decision_src})",
        }

    except json.JSONDecodeError as e:
        log.warning(f"Claude 回傳 JSON 解析失敗: {e} | raw={raw[:200]}")
        return None
    except Exception as e:
        log.warning(f"Claude API 呼叫失敗: {e}")
        return None


# ══════════════════════════════════════════════════════════
# 派單規則引擎（Claude AI + DB + Excel 三層）
# ══════════════════════════════════════════════════════════
def match_keyword(text: str, rules: list) -> dict | None:
    t = text.lower()
    for r in rules:
        if r["keyword"].lower() in t:
            return r
    return None

async def determine_assignment(summary: str, description: str = "",
                               requester: str = "") -> dict | None:
    """派單決策引擎（AI / 傳統模式皆含 CMDB 鑑別審查層）。

    AI 模式（Claude API Key 已設定）：
      ① Claude AI（信心 >= 0.85）→ 進入 CMDB 審查
           CMDB 審查通過 → 派單
           CMDB 審查不通過 → 掃 Description「Hi xxx」→ 二次人名 DB 比對
             二次信心 >= 0.85 → 採用（owner/assigned_to 用 DB 結果，inc_type 保留 Claude）
             二次信心 < 0.85  → 回傳 None（追蹤不派單）
      ② Claude 信心不足（< 0.85）或呼叫失敗 → 回傳 None

    傳統模式（無 Claude API Key）：
      ① SQLite 歷史比對（分數 >= db_similarity）
           分數 >= 0.75 → 進入 CMDB 鑑別審查（同上流程）
           分數 0.65-0.74 → 直接派單（略過 CMDB 審查）
      ② Excel 關鍵字規則
      ③ 預設值

    Returns:
        dict  → 有派單決策，可繼續派單
        None  → 任一層判定不通過，呼叫端負責追蹤記錄
    """
    combined = f"{summary} {description} {requester}"

    # ── Impact / Urgency / Priority 與 Due Date（所有模式共用）──────────
    # 週一(0)至週三(2)：Level 3 Minor/Medium，Due = +4 工作天
    # 週四(3)至週日(6)：Level 4 Routine/Low，  Due = +5 工作天
    _wd = datetime.now().weekday()
    if _wd <= 2:
        _impact, _urgency, _priority = "3 - Minor", "3 - Medium", "3 - Medium"
        _biz_days = 4
    else:
        _impact, _urgency, _priority = "4 - Routine", "4 - Low", "4 - Low"
        _biz_days = 5
    _due = calc_business_due_date(datetime.now(), _biz_days)
    _due_date = f"{_due.month}/{_due.day}/{_due.year}"
    log.info(
        f"Due Date: 今日星期{_wd + 1}(1=Mon) → +{_biz_days} 工作天 → {_due_date} | "
        f"Priority: {_priority}"
    )

    # ── 預先計算 Requester's Item 候選 + CMDB Direct 候選 ─────────────
    # 用 Excel 關鍵字規則從工單文字導出可能的 Requester's Item，
    # 再用該名稱查 CMDB 取得擁有者，作為「CMDB 直接派單」候選。
    # 此候選會在 AI/DB 兩條路徑作為 fallback 使用。
    _excel_ri_match = match_keyword(combined, RULES["requester_items"])
    _ri_candidate   = _excel_ri_match["item"] if _excel_ri_match else ""
    _excel_kw_match = match_keyword(combined, RULES["keyword_rules"])  # 預先取 Excel inc_type
    _cmdb_direct    = cmdb_direct_lookup(_ri_candidate)
    if _cmdb_direct:
        log.info(
            f"CMDB Direct 候選: Owner={_cmdb_direct['owner']} / "
            f"Assigned={_cmdb_direct['assigned_to']} "
            f"(CI={_cmdb_direct['requester_item']}, score={_cmdb_direct['_score']:.2f})"
        )

    def _fill_inc_type(base: dict) -> dict:
        """CMDB Direct 缺 Incident Type → 從 Excel 關鍵字補；都沒有則用安全預設值。"""
        if _excel_kw_match:
            base["inc_parent"] = _excel_kw_match["inc_parent"]
            base["inc_child"]  = _excel_kw_match["inc_child"]
            base["inc_item"]   = _excel_kw_match["inc_item"]
        else:
            base["inc_parent"] = "Service Request"
            base["inc_child"]  = "Infrastructure"
            base["inc_item"]   = "Standard Request"
        return base

    def _build_cmdb_direct_dispatch() -> dict:
        """組裝 CMDB Direct 派單結果（含 inc_type 補齊、impact/priority/due_date）。"""
        result = {
            "owner":          _cmdb_direct["owner"],
            "assigned_to":    _cmdb_direct["assigned_to"],
            "requester_item": _cmdb_direct["requester_item"],
            "impact":         _impact,
            "urgency":        _urgency,
            "priority":       _priority,
            "due_date":       _due_date,
            "_score":         _cmdb_direct["_score"],
            "_source":        _cmdb_direct["_source"],
        }
        return _fill_inc_type(result)

    # ══ AI 模式：Claude API Key 已設定 ════════════════════════════════
    if RULES.get("claude_api_key"):
        candidates = db_get_candidates(
            summary, description, requester,
            top_n=RULES.get("claude_candidates", 8)
        )
        ai_result = await claude_assign(summary, description, requester, candidates)

        if ai_result:
            # ── 三信號合一：Claude 已綜合 內容 + 歷史 + CMDB 做出決策 ──────────
            # 不再事後覆蓋。Claude 的 prompt 內每個歷史候選旁邊已標註該 CI 在
            # CMDB 中的擁有者，Claude 已能在單次推理中比較三個信號並選擇最佳
            # answer。系統職責只剩「執行 Claude 的決定 + 記錄三信號出處」。
            ai_conf     = ai_result.get("confidence", 0.85)
            ai_req_item = (ai_result.get("requester_item") or "").strip()
            ai_src      = ai_result.get("decision_source", "unknown")

            log.info(
                f"⚡ Claude 三信號決策: assigned={ai_result['assigned_to']} / "
                f"owner={ai_result['owner']} / CI={ai_req_item or '(無)'} "
                f"| 來源依據={ai_src} 信心={ai_conf:.2f}"
            )

            return {
                "owner":          ai_result["owner"],
                "assigned_to":    ai_result["assigned_to"],
                "inc_parent":     ai_result["inc_parent"],
                "inc_child":      ai_result["inc_child"],
                "inc_item":       ai_result["inc_item"],
                "requester_item": ai_req_item,
                "impact":         _impact,
                "urgency":        _urgency,
                "priority":       _priority,
                "due_date":       _due_date,
                "_score":         ai_conf,
                "_source":        ai_result["_source"],
            }
        else:
            # Claude 信心不足（< 0.85）或呼叫失敗 → 追蹤不派單
            log.warning(f"Claude 信心不足 (< 0.85)，工單僅追蹤不派單 | {summary[:50]}")
            return None

    # ══ 傳統模式：無 Claude API Key ═══════════════════════════════════
    hist = None
    db_score = 0.0

    # ① SQLite 歷史比對
    hist = db_find_similar(summary, description, requester)

    # ── CMDB Direct vs DB：取分數較高者作為主要派單來源 ──────────────
    # CMDB Direct 預設 _score=0.80，當 DB 命中分數低於此值時優先採用 CMDB Direct
    if _cmdb_direct and (not hist or _cmdb_direct["_score"] > hist.get("_score", 0.0)):
        log.info(
            f"CMDB Direct 勝過 DB（CMDB={_cmdb_direct['_score']:.2f} "
            f"vs DB={hist.get('_score', 0.0):.2f}），採用 CMDB 直接派單"
        )
        return _build_cmdb_direct_dispatch()

    if hist:
        db_score = hist.get("_score", 0.0)

        # ── CMDB 鑑別審查（0.75 <= DB 信心 <= 0.85 才觸發）─────────────
        # 信心 > 0.85：高信任，直接派單；信心 < 0.75：信心偏低，仍直接派單但不審查
        if db_score > 0.85:
            log.info(
                f"DB 信心 {db_score:.2f} > 0.85，跳過 CMDB 審查直接派單"
            )
        elif db_score >= 0.75:
            cmdb_ok, cmdb_reason = cmdb_validate(
                hist["assigned_to"],
                hist["owner"],
                hist.get("requester_item", ""),
            )
            log.info(
                f"CMDB 審查 (DB score={db_score:.2f}): {cmdb_reason}"
            )

            if not cmdb_ok:
                # 審查不通過 → 掃 Description 中的「Hi xxx」
                hi_name = extract_hi_name(description)
                log.info(
                    f"CMDB 審查不通過，掃描描述人名: "
                    f"{'「' + hi_name + '」' if hi_name else '（未找到）'}"
                )

                if hi_name:
                    # 二次人名 DB 比對（門檻 0.85）
                    second = db_find_by_name(hi_name, summary, description)
                    if second:
                        log.info(f"二次人名比對成功: {second['_source']}")
                        hist = second   # 採用二次比對結果
                    else:
                        log.warning(
                            f"⛔ 二次人名比對信心不足（名字: {hi_name}），"
                            f"工單僅追蹤記錄，不派單"
                        )
                        return None
                else:
                    log.warning(
                        "⛔ CMDB 審查不通過且描述中無法提取名字，"
                        "工單僅追蹤記錄，不派單"
                    )
                    return None
        # DB score < 0.75：信心偏低但仍達門檻，略過 CMDB 審查直接派單

        owner, assigned       = hist["owner"], hist["assigned_to"]
        inc_parent, inc_child = hist["inc_parent"], hist["inc_child"]
        inc_item              = hist["inc_item"]
        source                = hist["_source"]
    else:
        # ② Excel 關鍵字規則
        matched = match_keyword(combined, RULES["keyword_rules"])
        if matched:
            owner, assigned = matched["owner"], matched["assigned_to"]
            inc_parent, inc_child, inc_item = (
                matched["inc_parent"], matched["inc_child"], matched["inc_item"])
            source = "Excel關鍵字"
        else:
            # ③ 預設值
            log.warning(f"無匹配，使用預設 | {summary[:50]}")
            owner      = RULES["default_owner"]
            assigned   = RULES["default_assigned"]
            inc_parent = "Service Request"
            inc_child  = "Infrastructure" if owner.lower() in ["infra","help desk","helpdesk"] else "Application"
            inc_item   = "Standard Request"
            source     = "預設值"

    # Requester's Item：DB 歷史優先 → Excel → 空
    requester_item = (hist.get("requester_item") or "") if hist else ""
    if requester_item:
        log.info(f"  Requester's Item (DB ← {hist.get('_source','')}): {requester_item}")
    else:
        req = match_keyword(combined, RULES["requester_items"])
        if req:
            requester_item = req["item"]
            log.info(f"  Requester's Item (Excel): {requester_item}")
        else:
            log.debug("  Requester's Item: 無匹配，略過")

    # ── Impact / Urgency / Priority 與 Due Date 依派單當日星期決定 ───
    # 週一(0)至週三(2)：Level 3 Minor/Medium，Due = +4 工作天
    # 週四(3)至週日(6)：Level 4 Routine/Low，  Due = +5 工作天
    wd = datetime.now().weekday()   # 0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri, 5=Sat, 6=Sun
    if wd <= 2:
        impact, urgency, priority = "3 - Minor", "3 - Medium", "3 - Medium"
        biz_days = 4
    else:
        impact, urgency, priority = "4 - Routine", "4 - Low", "4 - Low"
        biz_days = 5

    due = calc_business_due_date(datetime.now(), biz_days)
    due_date = f"{due.month}/{due.day}/{due.year}"
    log.info(
        f"Due Date: 今日星期{wd+1}(0=Mon) → +{biz_days} 工作天 → {due_date} | "
        f"Priority: {priority}"
    )

    return {
        "owner": owner, "assigned_to": assigned,
        "inc_parent": inc_parent, "inc_child": inc_child, "inc_item": inc_item,
        "requester_item": requester_item,
        "impact": impact, "urgency": urgency, "priority": priority,
        "due_date": due_date, "_score": db_score, "_source": source,
    }


# ══════════════════════════════════════════════════════════
# UI 操作函式
# ══════════════════════════════════════════════════════════
async def select_native(page: Page, selector: str, value_map: dict, value_str: str, label: str):
    try:
        el = await page.query_selector(selector)
        if not el:
            base = selector.split("_ddl")[0].lstrip("#")
            el = await page.query_selector(f"select[id*='{base}']")
        if not el:
            log.warning(f"找不到 {label} <select>")
            return
        val = value_map.get(value_str, "")
        if val:
            await el.select_option(value=val)
            log.info(f"✓ {label}: {value_str}")
        else:
            await el.select_option(label=value_str)
            log.info(f"✓ {label}: {value_str} (by label)")
    except Exception as e:
        log.warning(f"select_native [{label}] 失敗: {e}")


def _extract_ad_account(name: str) -> str:
    """從顯示名稱中萃取 AD 帳號部分，提高 Person Chooser autocomplete 命中率。

    範例：
        'wen.hsieh 謝文譯'      → 'wen.hsieh'
        '謝文譯 wen.hsieh'      → 'wen.hsieh'
        'Alice Wang'            → 'Alice'  （沒明顯 AD pattern 時取首段）
        'Help Desk'             → 'Help Desk'（無空格時保留原樣）
        '謝文譯'                → '謝文譯' （純中文時保留）

    判斷規則：
      - 整串包含空格 → 嘗試找其中「ASCII 字母開頭、含 . _ - @ 等帳號字元」的段落
      - 找不到 AD pattern → 回傳第一個非空段落
      - 整串無空格 → 直接回傳
    """
    if not name:
        return ""
    name = name.strip()
    if " " not in name:
        return name
    parts = [p for p in name.split() if p]
    if not parts:
        return name
    # AD 帳號樣式：ASCII 字母開頭，且必須含有 . _ @ - 其中之一（避免把
    # 'Help Desk' 切成 'Help'、'Alice Wang' 切成 'Alice'）
    ad_pat = re.compile(r"^[A-Za-z][A-Za-z0-9]*[._@\-][A-Za-z0-9._@\-]+$")
    for p in parts:
        if ad_pat.match(p):
            return p
    # 沒有 AD 帳號 pattern → 保留原字串（給 Help Desk / Alice Wang 這種）
    return name


async def person_chooser_fill(page: Page, input_sel: str, value: str, label: str):
    """
    PersonChooser：input 隱藏時改用 JS focus + page.keyboard.type()
    - 空值直接略過
    - 固定 ID 找不到時，用 id 關鍵字做備援 wildcard 查詢
    - dropdown 出現後點擊第一筆，否則 Enter 送出
    """
    if not value or not value.strip():
        log.debug(f"[{label}] 值為空，略過")
        return

    # ── 萃取 AD 帳號（提高 autocomplete 命中率）─────────────────────
    # CMDB 存的是 "wen.hsieh 謝文譯" 這種顯示字串，整串打進 autocomplete
    # 通常 filter 不到。只用 AD 帳號 "wen.hsieh" 一定命中。
    type_value = _extract_ad_account(value)
    if type_value != value:
        log.debug(f"[{label}] 萃取 AD 帳號: '{value}' → '{type_value}'")

    try:
        el_id = input_sel.lstrip("#")
        # 從 id 提取關鍵字 (e.g. 'ItemOwner' or 'AssignedTo')
        id_parts = el_id.split("_")
        id_key   = id_parts[1] if len(id_parts) > 1 else el_id

        # Step 1: JS 強制可見 + focus（精確 ID → 備援 wildcard）
        found = await page.evaluate(f"""
            (function() {{
                var el = document.getElementById('{el_id}');
                if (!el) {{
                    var candidates = document.querySelectorAll('[id*="{id_key}"][id$="_I"]');
                    if (candidates.length > 0) el = candidates[0];
                }}
                if (!el) return false;
                el.style.cssText += '; display:block !important; visibility:visible !important; opacity:1 !important;';
                el.value = '';
                el.focus();
                return true;
            }})();
        """)
        if not found:
            log.warning(f"[{label}] 找不到 input (id={el_id})")
            return
        await page.wait_for_timeout(300)

        # Step 2: keyboard.type() 打字（只打 AD 帳號）
        await page.keyboard.type(type_value, delay=80)
        await page.wait_for_timeout(1800)

        # Step 3: 點擊 dropdown 第一筆「非空」項目（驗證 textContent 非空）
        #         避免點到 grid header 或其他空白 cell
        matched = await page.evaluate("""
            (function() {
                var selectors = [
                    '.dxeListBoxItem_Sunview',
                    '[class*="GridLookup"] td[class*="dxgv"]',
                    '.dxeListBoxItemRow td',
                    'td.dxgv'
                ];
                for (var s = 0; s < selectors.length; s++) {
                    var items = document.querySelectorAll(selectors[s]);
                    for (var i = 0; i < items.length; i++) {
                        var el  = items[i];
                        var txt = (el.textContent || '').trim();
                        // 必須可見 + textContent 非空（避免點到 header/空 cell）
                        if (el.offsetParent !== null && txt.length > 0) {
                            el.click();
                            return 'clicked: ' + txt.substring(0, 40);
                        }
                    }
                }
                return 'no dropdown';
            })();
        """)

        if matched == "no dropdown":
            # dropdown 沒出來 → 按 Enter 嘗試送出（ChangeGear 部分欄位接受）
            await page.keyboard.press("Enter")
            log.warning(
                f"⚠ {label}: '{type_value}' dropdown 未出現，改按 Enter（可能未真正選到人）"
            )
        else:
            log.info(f"✓ {label}: '{type_value}' → {matched}")

    except Exception as e:
        log.warning(f"person_chooser_fill [{label}] 失敗: {e}")


async def set_due_date(page: Page, date_str: str):
    try:
        inp = await page.wait_for_selector(SEL["due_date"], timeout=5000)
        await inp.evaluate("el => el.select()")
        await inp.fill(date_str)
        await inp.press("Tab")
        log.info(f"✓ Due Date: {date_str}")
    except Exception as e:
        log.warning(f"set_due_date 失敗: {e}")


async def set_incident_type(page: Page, level1: str, level2: str, level3: str):
    """Telerik RadTreeView — 全程 JS，繞過可見性限制"""
    try:
        # 開啟下拉
        await page.evaluate("""
            (function() {
                var els = document.querySelectorAll('[id]');
                for (var i = 0; i < els.length; i++) {
                    if (els[i].id.indexOf('IncidentRequestType') !== -1 &&
                        els[i].id.endsWith('_B-1')) {
                        els[i].click(); return;
                    }
                }
                var inp = document.querySelector('[id*="IncidentRequestType"][id$="_I"]');
                if (inp) { inp.click(); inp.focus(); }
            })();
        """)
        await page.wait_for_timeout(1000)

        async def js_expand(text: str) -> bool:
            result = await page.evaluate(f"""
                (function() {{
                    var spans = document.querySelectorAll('span.rtIn');
                    for (var i = 0; i < spans.length; i++) {{
                        if (spans[i].textContent.trim() === '{text}') {{
                            var li = spans[i].closest('li');
                            if (!li) {{ spans[i].click(); return 'clicked text (no li)'; }}
                            var plus = li.querySelector('span.rtPlus, span.rtExpand, .rtIcon');
                            if (plus) {{ plus.click(); return 'clicked plus'; }}
                            spans[i].click(); return 'clicked text';
                        }}
                    }}
                    return 'not found';
                }})();
            """)
            await page.wait_for_timeout(500)
            if result == "not found":
                log.warning(f"JS 找不到節點: {text}")
                return False
            return True

        async def js_click_leaf(text: str) -> bool:
            result = await page.evaluate(f"""
                (function() {{
                    var spans = document.querySelectorAll('span.rtIn');
                    for (var i = 0; i < spans.length; i++) {{
                        if (spans[i].textContent.trim() === '{text}') {{
                            spans[i].click(); return 'clicked';
                        }}
                    }}
                    return 'not found';
                }})();
            """)
            await page.wait_for_timeout(400)
            return result != "not found"

        ok1 = await js_expand(level1)
        ok2 = await js_expand(level2)
        ok3 = await js_click_leaf(level3)

        if ok1 and ok2 and ok3:
            log.info(f"✓ Incident Type: {level1} > {level2} > {level3}")
        else:
            log.warning(f"Incident Type 節點部分未找到 (l1={ok1} l2={ok2} l3={ok3})")

        # 收起下拉樹：用 Escape 鍵，避免 mouse.click 誤按 Action Bar 上的 Accept/Save 等按鈕
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(300)
    except Exception as e:
        log.warning(f"set_incident_type 失敗: {e}")


async def set_requester_item(page: Page, item_name: str):
    """
    Requester's Item popup 操作（已驗證完整流程）。

    流程（與人工操作完全一致）：
      1. click dropDownPopupButton（A tag）→ 開啟 "Select Item" popup
      2. 在 input 欄位輸入關鍵字 → 觸發 server-side 搜尋篩選
      3. 等待下拉出現 → 點選最佳匹配的 li.rcbItem
      4. Playwright click AddImageButton（INPUT type=image，需原生 click 帶座標）
      5. 驗證欄位值
      6. JS click CloseLinkButton 關閉 popup

    注意：
      - dropDownPopupButton 開啟 Select Item（含搜尋）；AddImpactedResouce_Button1 = Impact Analysis，勿用
      - 輸入關鍵字搜尋比點 Arrow 更可靠（可觸發 server 端分頁搜尋，結果更完整）
    """
    if not item_name:
        return
    try:
        combo_id  = SEL["ri_combo_id"]
        open_sel  = SEL["ri_open"]          # #{_RIRC}_dropDownPopupButton
        input_sel = f"#{combo_id}_Input"
        lb_sel    = f"#{combo_id}_listbox"
        add_sel   = f"#{_RIRC}_ASPxPopupControl1_AddImageButton"
        close_sel = f"#{_RIRC}_ASPxPopupControl1_CloseLinkButton"

        # 1. 開啟 popup
        await page.click(open_sel, timeout=5000)

        # 等 input 可見（popup 完整開啟）
        await page.wait_for_selector(input_sel, state="visible", timeout=8000)

        # 2. 輸入關鍵字觸發 server-side 搜尋
        await page.fill(input_sel, item_name)
        await page.wait_for_timeout(900)   # 等待 AJAX 回傳

        # 3. 從下拉清單點選最佳匹配的 li（精確優先，次選部分符合）
        matched = await page.evaluate(f"""
            (function() {{
                var lb = document.getElementById('{combo_id}_listbox');
                if (!lb) return 'listbox_not_found';

                var target = '{item_name}'.toLowerCase().trim();
                var lis = lb.querySelectorAll('li.rcbItem');
                if (!lis.length) return 'no_items';

                var exactEl = null, partialEl = null;
                for (var i = 0; i < lis.length; i++) {{
                    var t = lis[i].textContent.trim().toLowerCase();
                    if (t === target) {{ exactEl = lis[i]; break; }}
                    if (!partialEl && t.indexOf(target) !== -1) partialEl = lis[i];
                }}

                var el = exactEl || partialEl;
                if (!el) return 'no_match';
                el.click();
                return el.textContent.trim();
            }})();
        """)
        log.debug(f"Requester's Item 選取: {matched}")

        if matched in ('listbox_not_found', 'no_items', 'no_match'):
            log.info(f"Requester's Item 無匹配，略過 | 搜尋: {item_name}")
            try:
                await page.evaluate(f"document.getElementById('{_RIRC}_ASPxPopupControl1_CloseLinkButton').click();")
            except Exception:
                pass
            return

        await page.wait_for_timeout(300)

        # 4. 綠色 ⊕ Add — INPUT type="image"，需 Playwright 原生 click（帶座標）
        await page.click(add_sel, force=True, timeout=5000)
        await page.wait_for_timeout(800)

        # 5. 驗證欄位顯示值
        try:
            display_val = await page.evaluate(f"""
                var el = document.getElementById('{_RIRC}_dropDownTextBox_I');
                el ? el.value : '';
            """)
            if display_val:
                log.info(f"✓ Requester's Item: {display_val}")
            else:
                log.warning(f"Requester's Item Add 後欄位仍空 (matched={matched})")
        except Exception:
            log.debug("Requester's Item 驗證略過（頁面狀態變動）")

        # 6. 關閉 popup
        try:
            await page.evaluate(f"document.getElementById('{_RIRC}_ASPxPopupControl1_CloseLinkButton').click();")
            await page.wait_for_timeout(500)
        except Exception:
            pass

    except Exception as e:
        log.warning(f"set_requester_item 失敗: {e}")


async def click_save(page: Page) -> bool:
    """Save — JS click，繞過 tooltip 遮擋"""
    try:
        await page.mouse.move(0, 0)
        await page.wait_for_timeout(300)
        clicked = await page.evaluate("""
            (function() {
                var btn = document.getElementById('ActionBarControl1_Save_Button');
                if (!btn) return false;
                btn.click(); return true;
            })();
        """)
        if clicked:
            await page.wait_for_load_state("networkidle", timeout=20000)
            log.info("✓ Save 完成")
            return True
        log.warning("Save 按鈕不存在")
        return False
    except Exception as e:
        log.error(f"click_save 失敗: {e}")
        return False


async def dismiss_blocking_modals(page: Page) -> str:
    """掃描頁面上是否有「擋住派單」的對話框（Telerik RadWindow / ASPxPopup /
    OK-Cancel 確認框等），找到就嘗試點「OK / Yes / 確定」關閉。

    回傳：偵測到並嘗試處理的對話框文字（前 100 字），找不到則回傳空字串。
    """
    try:
        result = await page.evaluate("""
            (function() {
                // 常見的 ChangeGear / Telerik / DevExpress 對話框 selector
                var dialogSels = [
                    '.rwDialog:not([style*="display: none"])',
                    '.RadWindow:not([style*="display: none"])',
                    '.dxpc-content:not([style*="display: none"])',
                    '[role="dialog"]:not([style*="display: none"])',
                    '.modal.in', '.modal[style*="display: block"]'
                ];
                for (var s = 0; s < dialogSels.length; s++) {
                    var dialogs = document.querySelectorAll(dialogSels[s]);
                    for (var d = 0; d < dialogs.length; d++) {
                        var dlg = dialogs[d];
                        if (dlg.offsetParent === null) continue;
                        var dlgText = (dlg.innerText || '').trim();
                        // 嘗試找「OK / Yes / 確定 / 確認」按鈕
                        var btns = dlg.querySelectorAll('button, input[type="button"], a.rbButton');
                        for (var b = 0; b < btns.length; b++) {
                            var label = (btns[b].innerText || btns[b].value || '').trim();
                            if (/^(OK|Yes|確定|確認|Confirm)$/i.test(label)) {
                                btns[b].click();
                                return 'dismissed:' + dlgText.substring(0, 100);
                            }
                        }
                        return 'unhandled:' + dlgText.substring(0, 100);
                    }
                }
                return '';
            })();
        """)
        return result or ""
    except Exception as e:
        log.debug(f"dismiss_blocking_modals 失敗: {e}")
        return ""


async def verify_dispatch_success(page: Page) -> tuple[bool, str]:
    """
    Accept 之後驗證工單是否真的派出去了：
      1. Accept 按鈕是否還在（還在 = 狀態沒變 = 沒派出）
      2. Status 欄位內容是否還是 New
      3. 頁面是否有殘留對話框（會擋住儲存）
    """
    try:
        await page.wait_for_timeout(800)   # 讓 server postback 完成

        check = await page.evaluate("""
            (function() {
                var out = {accept_still: false, status_text: '', has_modal: false};
                out.accept_still = !!document.getElementById('ActionBarControl1_Accept_Button');
                // 找 Status field（ChangeGear 用 ddl 或 input）
                var statusEl = document.querySelector('[id*="Status"][id$="_I"]')
                            || document.querySelector('select[id*="Status"]');
                if (statusEl) {
                    out.status_text = statusEl.tagName === 'SELECT'
                        ? (statusEl.options[statusEl.selectedIndex] || {}).text || ''
                        : statusEl.value || '';
                }
                // 殘留 modal 檢查
                var modals = document.querySelectorAll('.rwDialog, .RadWindow, .dxpc-content, [role="dialog"]');
                for (var i = 0; i < modals.length; i++) {
                    if (modals[i].offsetParent !== null) { out.has_modal = true; break; }
                }
                return out;
            })();
        """)

        if check["has_modal"]:
            return False, "頁面殘留對話框未關閉"
        if check["accept_still"]:
            return False, f"Accept 按鈕仍存在（status={check['status_text'] or '?'})"
        # status 變了或找不到（找不到也可能是頁面 redirect 了 → 視為成功）
        return True, f"狀態={check['status_text'] or '(無 Status 欄位，可能已轉 view)'}"
    except Exception as e:
        return False, f"verify 失敗: {e}"


async def click_accept(page: Page) -> bool:
    """
    Accept — 在 New 狀態時點擊，同時儲存欄位並將票單轉為 In-Progress。
    Accept 按鈕不存在（票單已不是 New 狀態）→ 回傳 False，由呼叫端 fallback 到 Save。
    """
    for attempt in range(2):
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=8000)
            await page.wait_for_timeout(300)

            # 確認 Accept 按鈕是否存在
            exists = await page.evaluate("""
                !!document.getElementById('ActionBarControl1_Accept_Button')
            """)
            if not exists:
                log.debug("Accept 按鈕不存在（票單非 New 狀態）")
                return False   # ← 讓呼叫端知道要 fallback

            # 存在 → 點擊
            await page.evaluate("""
                document.getElementById('ActionBarControl1_Accept_Button').click();
            """)
            await page.wait_for_load_state("networkidle", timeout=10000)

            # 點擊後檢查並關閉「Are you sure?」之類的確認對話框
            modal = await dismiss_blocking_modals(page)
            if modal:
                log.info(f"Accept 後對話框: {modal[:80]}")
                # 對話框點掉之後再等網路完成
                try:
                    await page.wait_for_load_state("networkidle", timeout=5000)
                except Exception:
                    pass

            log.info("✓ Accept 完成")
            return True

        except Exception as e:
            if "Execution context was destroyed" in str(e) and attempt == 0:
                log.debug("Accept context 被銷毀，等候後重試...")
                await page.wait_for_timeout(1500)
            else:
                log.warning(f"click_accept 失敗（第{attempt+1}次）: {e}")
                return False
    return False


# ══════════════════════════════════════════════════════════
# ChangeGear Bot
# ══════════════════════════════════════════════════════════
class ChangeGearBot:
    def __init__(self):
        self.playwright = None
        self.browser    = None
        self.context    = None
        self.page       = None

    async def start(self):
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=RULES["headless"])
        self.context = await self.browser.new_context(
            http_credentials={"username": RULES["ad_account"], "password": RULES["ad_password"]}
        )
        self.page = await self.context.new_page()
        log.info("瀏覽器已啟動")

    async def stop(self):
        if self.browser:    await self.browser.close()
        if self.playwright: await self.playwright.stop()

    async def go_to_list(self) -> bool:
        url = (f"{RULES['base_url']}/MainUI/Common/Modules/BaseModule.aspx"
               "?ModuleName=Incident&view=All%20New%20Incidents&text=New%20Incidents")
        for attempt in range(2):
            try:
                await self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
                await self.page.wait_for_load_state("networkidle", timeout=15000)
                return True
            except Exception as e:
                if "ERR_ABORTED" in str(e) or "net::" in str(e):
                    log.warning(f"列表頁被中止，重試 (attempt {attempt+1})")
                    await self.page.wait_for_timeout(2000)
                else:
                    log.error(f"載入列表頁失敗: {e}")
                    return False
        return False

    async def get_unassigned_tickets(self) -> list:
        """解析 All New Incidents 列表，回傳未派單工單。

        All New Incidents view 欄位索引（已驗證）：
          _1  = Item ID (label)
          _5  = Requester (.GridDataItem)  ← 寄件者
          _9  = Summary (.GridDataItem)    ← 信件主旨
          _12 = Assigned To (.GridDataItem)
        """
        tickets = []
        try:
            rows = await self.page.query_selector_all("tr[id*='DXDataRow']")
            for row in rows:
                ondblclick = await row.get_attribute("ondblclick") or ""
                m = re.search(r"OnGetRowValues\('(\d+)'\)", ondblclick)
                if not m:
                    continue
                oid = m.group(1)

                item_el    = await row.query_selector("td[id$='_1'] label")
                item_id    = (await item_el.inner_text()).strip() if item_el else ""

                req_el     = await row.query_selector("td[id$='_5'] .GridDataItem")
                requester  = (await req_el.inner_text()).strip() if req_el else ""

                summary_el = await row.query_selector("td[id$='_9'] .GridDataItem")
                summary    = (await summary_el.inner_text()).strip() if summary_el else ""

                assign_el  = await row.query_selector("td[id$='_12'] .GridDataItem")
                assign_to  = (await assign_el.inner_text()).strip() if assign_el else ""

                if not assign_to or assign_to == "\xa0":
                    tickets.append({
                        "oid": oid, "item_id": item_id,
                        "summary": summary, "requester": requester
                    })
                    log.info(f"未派單: {item_id} | From: {requester} | {summary[:40]}")
        except Exception as e:
            log.error(f"解析列表失敗: {e}")
        return tickets

    async def open_ticket(self, oid: str) -> bool:
        url = (f"{RULES['base_url']}/MainUI/ServiceDesk/SDItemEditPanel.aspx"
               f"?boundtable=IIncidentRequest&CloseOnPerformAction=false"
               f"&ID={oid}&windowWidth=1050&refreshOnClose=true")
        try:
            await self.page.goto(url, wait_until="networkidle", timeout=30000)
            return True
        except Exception as e:
            log.error(f"開啟工單 {oid} 失敗: {e}")
            return False

    async def get_description(self) -> str:
        """讀取工單描述（= 信件內文）。

        selector 已驗證: [id*='Description'][id*='_divComment']
        （DevExpress comment 區塊，包含完整信件內容）
        """
        try:
            el = await self.page.query_selector("[id*='Description'][id*='_divComment']")
            return (await el.inner_text()).strip()[:1000] if el else ""
        except Exception:
            return ""

    async def read_current_assignment(self) -> dict:
        """讀取目前工單詳細頁的所有派單欄位現值（用於比對 bot 原本填的值）。

        已驗證 selector：
          Owner       : [id*='ItemOwner'][id$='_I']
          Assigned To : [id*='AssignedTo'][id$='_I']
          IncType     : [id*='IncidentRequestType'][id$='_I']
          Req Item    : [id*='ImpactedResourcesPIT'][id*='dropDownTextBox_I']
        """
        async def val(sel):
            try:
                el = await self.page.query_selector(sel)
                return (await el.input_value()).strip() if el else ""
            except Exception:
                return ""

        inc_raw   = await val("[id*='IncidentRequestType'][id$='_I']")
        parts     = [p.strip() for p in re.split(r"[-:>]", inc_raw) if p.strip()]
        inc_parent = parts[0] if len(parts) > 0 else ""
        inc_child  = parts[1] if len(parts) > 1 else ""
        inc_item   = parts[2] if len(parts) > 2 else ""

        return {
            "owner":       await val("[id*='ItemOwner'][id$='_I']"),
            "assigned_to": await val("[id*='AssignedTo'][id$='_I']"),
            "inc_parent":  inc_parent,
            "inc_child":   inc_child,
            "inc_item":    inc_item,
            "req_item":    await val(f"[id*='ImpactedResourcesPIT'][id*='dropDownTextBox_I']"),
        }

    @staticmethod
    def _is_real_correction(field: str, db_v: str, cur_v: str) -> bool:
        """判斷某欄位的「DB 值 → 現值」是否算「真修正」。

        排除以下三種「假修正」：
          1. Req Item 多值補資料：bot 填了 A，人工新增 A;B → 不算修正
          2. Owner / Assigned 顯示變體：bot=`wen.hsieh`，現值=`wen.hsieh 謝文譯`
             → 子字串相符，視為同一人
          3. Inc 欄位被動補資料：bot 留空白，人工後補 → 不算修正
        """
        db_v  = (db_v or "").strip()
        cur_v = (cur_v or "").strip()

        # 現值空 → 頁面未載入或欄位被清除，保守不算
        if not cur_v:
            return False
        if db_v == cur_v:
            return False

        # ── Req Item: 多值（分號或逗號分隔）──────────────────
        # bot 值若是現值的子集（例：bot=A，現值=A;B）→ 人工補了第二項，不算修正
        if field == "req_item":
            if not db_v:
                return False                 # bot 留空 → 人工後補 → 不算
            db_items  = {x.strip().lower() for x in re.split(r"[;,]", db_v)  if x.strip()}
            cur_items = {x.strip().lower() for x in re.split(r"[;,]", cur_v) if x.strip()}
            if db_items.issubset(cur_items):
                return False                  # 只是新增項目
            return True

        # ── Owner / Assigned: 子字串比對（顯示變體）─────────
        # 例：bot=`wen.hsieh`、現值=`wen.hsieh 謝文譯` → 同一人
        if field in ("owner", "assigned_to"):
            dl = db_v.lower()
            cl = cur_v.lower()
            if dl in cl or cl in dl:
                return False
            return True

        # ── Inc Parent / Child / Item: bot 留空白 + 人工補 → 不算 ──
        if not db_v:
            return False
        return True


    async def scan_and_learn_corrections(self):
        """掃描 bot 近期派單，偵測人工修正並回寫 DB 學習。

        流程：
          1. 從 DB 取出最近 7 天由 bot 派單且尚未標為已修正的記錄
          2. 開啟每張工單，讀取現值
          3. 若有任一欄位與 DB 記錄不同 → 視為人工修正
          4. 以修正後的值更新 DB（corrected=1），供下次比對使用
        """
        log.info("── 修正學習掃描開始 ──")
        try:
            conn = sqlite3.connect(DB_PATH)
            rows = conn.execute("""
                SELECT ticket_id, oid, owner, assigned_to,
                       inc_parent, inc_child, inc_item, req_item
                FROM   assignments
                WHERE  bot_assigned = 1
                AND    created_at  >= datetime('now', '-7 days')
                ORDER  BY created_at DESC
            """).fetchall()
            conn.close()
        except Exception as e:
            log.warning(f"修正掃描 DB 查詢失敗: {e}")
            return

        learned = 0
        if not rows:
            log.info("近 7 天無 bot 派單記錄（bot_assigned=1）需要檢查")
        else:
            log.info(f"檢查 {len(rows)} 張近期 bot 派單...")

        for row in rows:
            ticket_id, oid = row[0], row[1]
            db_vals = {
                "owner": row[2], "assigned_to": row[3],
                "inc_parent": row[4], "inc_child": row[5],
                "inc_item": row[6], "req_item": row[7],
            }

            if not oid:
                log.debug(f"{ticket_id}: 無 oid，跳過")
                continue

            try:
                url = (f"{RULES['base_url']}/MainUI/ServiceDesk/SDItemEditPanel.aspx"
                       f"?boundtable=IIncidentRequest&CloseOnPerformAction=false"
                       f"&ID={oid}&windowWidth=1050&refreshOnClose=true")
                await self.page.goto(url, wait_until="domcontentloaded", timeout=20000)
                await self.page.wait_for_timeout(1500)

                cur = await self.read_current_assignment()

                # 比較有無差異 — 用新版 helper 過濾掉「補資料 / 顯示變體」
                diffs = {}
                soft_diffs = {}      # 有變但不算修正（補資料等）
                for k in ("owner", "assigned_to", "inc_parent", "inc_child",
                          "inc_item", "req_item"):
                    db_v  = (db_vals[k] or "").strip()
                    cur_v = (cur[k] or "").strip()
                    if cur_v and cur_v != db_v:
                        if self._is_real_correction(k, db_v, cur_v):
                            diffs[k] = {"was": db_v, "now": cur_v}
                        else:
                            soft_diffs[k] = {"was": db_v, "now": cur_v}

                if not diffs:
                    if soft_diffs:
                        log.debug(f"{ticket_id}: 軟修正（不算）: {soft_diffs}")
                    else:
                        log.debug(f"{ticket_id}: 無修正")
                    continue

                # 有差異 → 人工修正了，更新 DB
                log.info(f"⚡ {ticket_id} 偵測到修正: {diffs}")
                conn = sqlite3.connect(DB_PATH)
                conn.execute("""
                    UPDATE assignments SET
                        owner       = ?,
                        assigned_to = ?,
                        inc_parent  = ?,
                        inc_child   = ?,
                        inc_item    = ?,
                        req_item    = ?,
                        corrected   = 1,
                        updated_at  = CURRENT_TIMESTAMP
                    WHERE ticket_id = ?
                """, (
                    cur["owner"]       or db_vals["owner"],
                    cur["assigned_to"] or db_vals["assigned_to"],
                    cur["inc_parent"]  or db_vals["inc_parent"],
                    cur["inc_child"]   or db_vals["inc_child"],
                    cur["inc_item"]    or db_vals["inc_item"],
                    cur["req_item"]    or db_vals["req_item"],
                    ticket_id,
                ))
                conn.commit()
                conn.close()
                learned += 1

            except Exception as e:
                log.debug(f"{ticket_id} 修正掃描失敗: {e}")
                continue

        log.info(f"── bot 派單修正學習完成，更新 {learned} 筆 ──")

        # ── 追蹤工單學習（bot_assigned=2）──────────────────────────────
        # 這些工單是「偵測到但非 Help Desk / 無法判斷」而未派單的單據。
        # 若人工已填入資料，則學習並提升至 corrected=1，供未來比對使用。
        log.info("── 追蹤工單學習掃描開始 ──")
        try:
            conn = sqlite3.connect(DB_PATH)
            tracked = conn.execute("""
                SELECT ticket_id, oid
                FROM   assignments
                WHERE  bot_assigned = 2
                AND    created_at  >= datetime('now', '-14 days')
                ORDER  BY created_at DESC
            """).fetchall()
            conn.close()
        except Exception as e:
            log.warning(f"追蹤工單 DB 查詢失敗: {e}")
            tracked = []

        if not tracked:
            log.info("無追蹤工單需要檢查")
        else:
            log.info(f"檢查 {len(tracked)} 張追蹤工單...")
            track_learned = 0
            for ticket_id, oid in tracked:
                if not oid:
                    continue
                try:
                    url = (f"{RULES['base_url']}/MainUI/ServiceDesk/SDItemEditPanel.aspx"
                           f"?boundtable=IIncidentRequest&CloseOnPerformAction=false"
                           f"&ID={oid}&windowWidth=1050&refreshOnClose=true")
                    await self.page.goto(url, wait_until="domcontentloaded", timeout=20000)
                    await self.page.wait_for_timeout(1500)

                    cur = await self.read_current_assignment()

                    # 只要 owner 或 assigned_to 有值，代表已有人工填入
                    if not (cur.get("owner") or cur.get("assigned_to")):
                        log.debug(f"{ticket_id}: 追蹤中，尚無人工填入")
                        continue

                    log.info(
                        f"⚡ {ticket_id} 追蹤工單已被人工填入 → 學習 | "
                        f"Owner={cur['owner']} / Assigned={cur['assigned_to']} / "
                        f"Type={cur['inc_parent']}>{cur['inc_child']}>{cur['inc_item']}"
                    )
                    conn = sqlite3.connect(DB_PATH)
                    conn.execute("""
                        UPDATE assignments SET
                            owner       = ?,
                            assigned_to = ?,
                            inc_parent  = ?,
                            inc_child   = ?,
                            inc_item    = ?,
                            req_item    = ?,
                            bot_assigned = 0,
                            corrected   = 1,
                            updated_at  = CURRENT_TIMESTAMP
                        WHERE ticket_id = ?
                    """, (
                        cur["owner"], cur["assigned_to"],
                        cur["inc_parent"], cur["inc_child"], cur["inc_item"],
                        cur["req_item"], ticket_id,
                    ))
                    conn.commit()
                    conn.close()
                    track_learned += 1

                except Exception as e:
                    log.debug(f"{ticket_id} 追蹤學習掃描失敗: {e}")
                    continue

            log.info(f"── 追蹤工單學習完成，更新 {track_learned} 筆 ──")

    async def get_requester(self) -> str:
        """讀取工單 Requester（= 寄件者）。

        selector 已驗證: [id*='Requester'][id$='_I']
        → DynamicLayoutControl1_Requester_PersonChooser_GridLookupPC_I
        value 格式: "ryan.huang Ryan Huang"
        """
        try:
            el = await self.page.query_selector("[id*='Requester'][id$='_I']")
            return (await el.input_value()).strip() if el else ""
        except Exception:
            return ""

    async def assign_ticket(self, oid: str, item_id: str, summary: str,
                             requester: str = "") -> bool:
        log.info(f"── 開始派單: {item_id} (Oid={oid}) ──")
        if not await self.open_ticket(oid):
            return False

        # 從工單詳細頁讀取信件內文與 Requester（清單頁已有 requester，詳細頁再確認）
        desc = await self.get_description()
        req_detail = await self.get_requester()
        # 詳細頁的 requester 優先（更完整），清單頁的備用
        requester = req_detail or requester

        a = await determine_assignment(summary, desc, requester)

        # ── 判定不通過（Claude 信心不足 / CMDB 審查未過）→ 僅追蹤，不派單 ──
        if a is None:
            log.info(
                f"⏭ {item_id} 派單判定未通過"
                f"（Claude 信心不足 或 CMDB 審查未通過 或 二次比對信心不足），"
                f"僅追蹤記錄不派單"
            )
            db_track_ticket(item_id, oid, summary, desc, requester)
            return True

        log.info(
            f"派單決策({a['_source']}) | Requester={requester} | "
            f"Owner={a['owner']} | Assigned={a['assigned_to']} | "
            f"Type={a['inc_parent']}>{a['inc_child']}>{a['inc_item']} | "
            f"Priority={a['priority']} | Due={a['due_date']} | ReqItem={a['requester_item']}"
        )

        # ── Help Desk 過濾：非 Help Desk 工單只追蹤不派單 ───────────────
        # 例外：信心分數 >= 0.75 時，信任比對結果，跳過此過濾直接派單
        if not is_helpdesk_owner(a.get("owner", "")):
            _a_score = a.get("_score", 0.0)
            if _a_score >= 0.75:
                log.info(
                    f"⚡ {item_id} 非 Help Desk 工單（Owner={a.get('owner','（無）')}）"
                    f"，但信心分數 {_a_score:.2f} >= 0.75，跳過 Help Desk 過濾繼續派單"
                )
            else:
                log.info(
                    f"⏭ {item_id} 非 Help Desk 工單（Owner={a.get('owner','（無）')}），"
                    f"信心分數 {_a_score:.2f} < 0.75，僅追蹤記錄，待學習"
                )
                db_track_ticket(item_id, oid, summary, desc, requester)
                return True

        try:
            # ① Incident Type（選取後等待 AJAX postback 完成，DOM 才穩定）
            await set_incident_type(self.page, a["inc_parent"], a["inc_child"], a["inc_item"])
            try:
                await self.page.wait_for_load_state("networkidle", timeout=6000)
            except Exception:
                await self.page.wait_for_timeout(1500)

            # ② Owner（等 Owner input 出現再填）
            try:
                await self.page.wait_for_selector(SEL["owner_input"], timeout=5000)
            except Exception:
                pass
            await person_chooser_fill(self.page, SEL["owner_input"], a["owner"], "Owner")
            await self.page.wait_for_timeout(300)

            # ③ Assigned To（等 Assigned To input 出現再填）
            try:
                await self.page.wait_for_selector(SEL["assigned_input"], timeout=5000)
            except Exception:
                pass
            await person_chooser_fill(self.page, SEL["assigned_input"], a["assigned_to"], "Assigned To")
            await self.page.wait_for_timeout(300)

            # ④ Impact
            await select_native(self.page, SEL["impact"], IMPACT_MAP, a["impact"], "Impact")
            await self.page.wait_for_timeout(200)

            # ⑤ Urgency
            await select_native(self.page, SEL["urgency"], URGENCY_MAP, a["urgency"], "Urgency")
            await self.page.wait_for_timeout(200)

            # ⑥ Priority
            await select_native(self.page, SEL["priority"], PRIORITY_MAP, a["priority"], "Priority")
            await self.page.wait_for_timeout(200)

            # ⑦ Due Date
            await set_due_date(self.page, a["due_date"])
            await self.page.wait_for_timeout(200)

            # ⑧ Requester's Item（popup 流程）
            if a["requester_item"]:
                await set_requester_item(self.page, a["requester_item"])
                await self.page.wait_for_timeout(200)

            # ⑨ Accept（主要動作：儲存所有欄位 + 將票單從 New 轉 In-Progress）
            #    Accept 必須在 New 狀態時點擊，Save 會讓票單進入 In-Progress 導致 Accept 消失
            #    因此改為 Accept 優先，Accept 失敗才 fallback 到 Save
            await self.page.wait_for_timeout(500)
            accepted = await click_accept(self.page)

            if not accepted:
                # Accept 不可用（非 New 狀態）→ fallback 到 Save
                log.info("Accept 不可用，改用 Save")
                saved = await click_save(self.page)
                if not saved:
                    await self.page.screenshot(path=f"debug_{item_id}.png")
                    log.warning(f"⚠ Save 失敗，截圖: debug_{item_id}.png")
                    return False
            else:
                # Accept 「點擊」成功，但要驗證狀態真的變了
                # 之前遇過：log 寫 Accept 完成，但工單狀態仍 New（modal 沒關 / 欄位被拒）
                ok, why = await verify_dispatch_success(self.page)
                if not ok:
                    log.warning(f"⚠ {item_id} Accept 點擊但狀態未變：{why}，補打 Save")
                    # 先再清一次可能殘留的對話框
                    modal_msg = await dismiss_blocking_modals(self.page)
                    if modal_msg:
                        log.info(f"  清理對話框: {modal_msg[:80]}")
                    # 試 Save 補救
                    saved = await click_save(self.page)
                    if not saved:
                        await self.page.screenshot(path=f"debug_accept_fail_{item_id}.png")
                        log.warning(
                            f"⚠ {item_id} Save 補救也失敗，截圖: debug_accept_fail_{item_id}.png"
                        )
                        return False
                    # 補救後再驗證一次
                    ok2, why2 = await verify_dispatch_success(self.page)
                    if not ok2:
                        await self.page.screenshot(path=f"debug_save_fail_{item_id}.png")
                        log.warning(
                            f"⚠ {item_id} 補 Save 後仍未派出：{why2}，"
                            f"截圖: debug_save_fail_{item_id}.png"
                        )
                        return False
                    log.info(f"  ✓ Save 補救成功：{why2}")
                else:
                    log.debug(f"  Accept 驗證通過：{why}")

            # ⑪ 儲存到 SQLite 歷史 DB（含 oid + requester）
            db_save(item_id, oid, summary, desc, requester, a)

            log.info(f"✅ 完成: {item_id} → {a['assigned_to']}")
            return True

        except Exception as e:
            log.error(f"❌ {item_id} 失敗: {e}", exc_info=True)
            try:
                await self.page.screenshot(path=f"error_{item_id}.png")
            except Exception:
                pass
            return False

    async def run_scan(self):
        """主掃描：找未派單工單 → 派單"""
        log.info(f"===== 掃描開始 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} =====")
        try:
            if not await self.go_to_list():
                return
            tickets = await self.get_unassigned_tickets()
            log.info(f"找到 {len(tickets)} 張未派單工單")

            success = fail = 0
            for t in tickets:
                ok = await self.assign_ticket(
                    t["oid"], t["item_id"], t["summary"], t.get("requester", "")
                )
                if ok: success += 1
                else:  fail += 1
                await self.go_to_list()

            log.info(f"本次完成 — 成功:{success} 失敗:{fail}")
        except Exception as e:
            log.error(f"掃描異常: {e}", exc_info=True)
        log.info("===== 掃描結束 =====\n")

    async def run_correction_scan(self):
        """修正學習掃描：偵測人工修正 → 更新 DB（每小時執行一次）"""
        try:
            await self.scan_and_learn_corrections()
            await self.go_to_list()   # 掃完回到列表備用
        except Exception as e:
            log.error(f"修正掃描異常: {e}", exc_info=True)

    async def run_weekly_reflection(self):
        """每週反思：把過去 7 天的 feedback + 修正餵給 Claude，歸納派單原則
        寫入 learned_principles.md。系統下次派單時會自動載入。
        """
        log.info("════ 每週反思開始 ════")
        try:
            import subprocess
            proc = await asyncio.create_subprocess_exec(
                sys.executable, "reflect.py", "--days", "7",
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await proc.communicate()
            output = (stdout or b"").decode("utf-8", errors="replace")
            err    = (stderr or b"").decode("utf-8", errors="replace")
            for line in output.splitlines():
                if line.strip():
                    log.info(f"  [reflect] {line}")
            if proc.returncode != 0:
                log.warning(f"反思執行碼非 0: {proc.returncode} | stderr={err[-300:]}")
            else:
                log.info("════ 每週反思完成 ════")
        except Exception as e:
            log.error(f"每週反思異常: {e}", exc_info=True)


# ══════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════
async def main():
    init_db()
    log.info("ChangeGear 自動派單程式 v6 啟動")
    log.info(f"關鍵字規則: {len(RULES['keyword_rules'])} 條 | "
             f"Requester's Item: {len(RULES['requester_items'])} 條 | "
             f"DB相似度門檻: {RULES['db_similarity']}")
    log.info(f"掃描間隔: 每 {RULES['scan_interval']} 分鐘")

    bot = ChangeGearBot()
    await bot.start()

    scheduler = AsyncIOScheduler()
    # 派單掃描：每 N 分鐘（預設 30 分鐘）
    scheduler.add_job(
        bot.run_scan, "interval",
        minutes=RULES["scan_interval"],
        next_run_time=datetime.now(),
    )
    # 修正學習掃描：每 60 分鐘（首次延遲 5 分鐘後啟動）
    scheduler.add_job(
        bot.run_correction_scan, "interval",
        minutes=60,
        next_run_time=datetime.now() + timedelta(minutes=5),
    )
    # 每週反思：每週一 06:00 自動跑 reflect.py 歸納派單原則
    scheduler.add_job(
        bot.run_weekly_reflection, "cron",
        day_of_week="mon", hour=6, minute=0,
    )
    log.info("排程：派單掃描每 %d 分鐘 / 修正掃描每 60 分鐘 / 每週反思週一 06:00",
             RULES["scan_interval"])
    scheduler.start()

    try:
        await asyncio.Event().wait()
    except (KeyboardInterrupt, SystemExit):
        log.info("停止中...")
    finally:
        scheduler.shutdown()
        await bot.stop()
        log.info("程式已停止")


if __name__ == "__main__":
    asyncio.run(main())
